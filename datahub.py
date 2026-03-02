"""
datahub.py  —  v5.0
---------------------
L3 — DataHub: Google Sheet / xlsx → PrimaryData

V1 SHEET CONTRACT (finalised):
  Meta rows           — TICKER SYMBOL, COMPANY NAME, EXCHANGE, INDUSTRY
  Quarterly Results   — periodic P&L (quarterly or semi-annual for SME)
  Annual P&L          — source of truth for full-year income
  Balance Sheet       — annual, includes debt breakdown + working capital items
  Cashflow statement  — annual, includes CFO / CFI / CFF
  Ratios              — annual working capital ratios + ROCE
  Shareholding Pattern— quarterly: Promoters, FIIs, DIIs, Government, Public
  Links               — transcript + annual report PDF URLs

DESIGN RULES:
  1.  Annual P&L     → primary source for IncomeStatement
  2.  TTM column     → stored in ttm_* fields, excluded from annual series
  3.  Periodic P&L   → last 4 periods stored in recent_periods for trend analysis
  4.  Frequency      → auto-detected from date gaps (quarterly / semi-annual)
  5.  All annual series use get_floats() not sparse_series() — None preserved
  6.  Label matching → normalised (_norm) so +/- suffix variants always resolve
  7.  If source changes → only this file changes. Agents never change.
"""

import os
import datetime
import re
from typing import Optional, Any

from primary_data import (
    PrimaryData, CompanyMeta, IncomeStatement, BalanceSheet,
    CashflowStatement, ValuationInputs, RiskMeta, Exchange,
    ShareholdingData, OperatingMetrics,
)


# ─── Cell helpers ──────────────────────────────────────────────────────────────

def _f(val: Any) -> Optional[float]:
    """Safe float. Handles None, '', '176', '17.10%', '1,234'."""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').replace('%', '').strip())
    except (ValueError, TypeError):
        return None


def _pct(val: Any) -> Optional[float]:
    """
    Convert to percentage float.
      0.18   → 18.0   (decimal stored in sheet)
      17.10% → 17.1   (already a percentage string)
    """
    if val is None or val == '':
        return None
    s = str(val).strip()
    is_pct_string = s.endswith('%')
    v = _f(val)
    if v is None:
        return None
    if is_pct_string:
        return v
    return v * 100 if abs(v) < 2.0 else v


def _s(val: Any) -> Optional[str]:
    """Safe string — returns None for empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _date_label(d: Any) -> str:
    """datetime → 'Mar-2024'."""
    if isinstance(d, datetime.datetime):
        return d.strftime('%b-%Y')
    return str(d)


def _norm(label: str) -> str:
    """
    Normalise label for fuzzy matching.
    Strips trailing +/- and whitespace so that:
      'Borrowings +'  == 'Borrowings -'  == 'Borrowings '  == 'Borrowings'
      'Cash from Operating Activity +' matches 'Cash from Operating Ac'
    """
    if not label:
        return ''
    return label.strip().rstrip(' +-').lower()


# ─── Date string parser ────────────────────────────────────────────────────────

MONTH_MAP = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
    'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
    'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}


def _parse_date(cell: Any) -> Any:
    """
    Convert date strings → datetime.
    Handles: 'Dec 2022', 'Dec-2022', '2022-12-01', '2022-12-01 00:00:00'
    Returns original value if not a recognisable date.
    """
    if isinstance(cell, datetime.datetime):
        return cell
    if not cell or not str(cell).strip():
        return cell
    s = str(cell).strip()

    # 'Dec 2022' or 'Dec-2022'
    m = re.match(r'^([A-Za-z]{3})[\s\-](\d{4})$', s)
    if m:
        mon = MONTH_MAP.get(m.group(1).lower())
        yr  = int(m.group(2))
        if mon:
            return datetime.datetime(yr, mon, 1)

    # '2022-12-01' or '2022-12-01 00:00:00'
    m2 = re.match(r'^(\d{4}-\d{2}-\d{2})', s)
    if m2:
        try:
            return datetime.datetime.strptime(m2.group(1), '%Y-%m-%d')
        except ValueError:
            pass

    return cell


# ─── Reporting frequency detection ────────────────────────────────────────────

def _detect_frequency(dates: list) -> str:
    """
    Infer reporting frequency from datetime list.
    Returns: 'quarterly' | 'semi-annual' | 'annual' | 'unknown'
    """
    if len(dates) < 2:
        return 'unknown'
    gaps = []
    for i in range(1, len(dates)):
        if isinstance(dates[i], datetime.datetime) and isinstance(dates[i-1], datetime.datetime):
            gaps.append((dates[i] - dates[i-1]).days)
    if not gaps:
        return 'unknown'
    gaps.sort()
    median = gaps[len(gaps) // 2]
    if median < 120:
        return 'quarterly'
    elif median < 270:
        return 'semi-annual'
    return 'annual'


# ─── SheetSection ──────────────────────────────────────────────────────────────

class SheetSection:
    """
    One named section of the sheet.

    label_col  — column index where row labels live
                 0 for both gspread and V1 xlsx (labels always in col A)
    data_start — first column with data values (always label_col + 1)
    ttm_col    — column index of TTM value if present (not a datetime), else None
    """

    def __init__(self, section_name: str, date_row: list, data_rows: list,
                 label_col: int = 0, data_start: int = 1):
        self.name       = section_name
        self.label_col  = label_col
        self.data_start = data_start

        # Find datetime columns and TTM column separately
        self.date_cols: list = []
        self.ttm_col: Optional[int] = None

        for i, v in enumerate(date_row):
            if i < data_start:
                continue
            if isinstance(v, datetime.datetime):
                self.date_cols.append(i)
            elif _s(v) and _s(v).upper() == 'TTM':
                self.ttm_col = i

        self.dates: list     = [_date_label(date_row[i]) for i in self.date_cols]
        self.raw_dates: list = [date_row[i] for i in self.date_cols
                                if isinstance(date_row[i], datetime.datetime)]
        self.frequency: str  = _detect_frequency(self.raw_dates)

        # Build exact and normalised label lookups
        self._rows:      dict = {}
        self._norm_rows: dict = {}
        self._ttm:       dict = {}  # label → TTM value

        for row in data_rows:
            row = list(row) + [None] * 15
            label = _s(row[label_col])
            if not label:
                continue
            vals = [row[i] if i < len(row) else None for i in self.date_cols]
            self._rows[label] = vals
            nk = _norm(label)
            if nk and nk not in self._norm_rows:
                self._norm_rows[nk] = vals
            # TTM value
            if self.ttm_col is not None:
                self._ttm[label]  = row[self.ttm_col]
                self._ttm[nk]     = row[self.ttm_col]

    # ── Core accessors ──

    def get(self, label: str) -> list:
        """Raw values aligned to date_cols. Normalised fallback on miss."""
        if label in self._rows:
            return self._rows[label]
        return self._norm_rows.get(_norm(label), [])

    def get_floats(self, label: str) -> list:
        """Float values — None preserved for missing/blank cells."""
        return [_f(v) for v in self.get(label)]

    def get_pct(self, label: str) -> list:
        """Percentage values — None preserved."""
        return [_pct(v) for v in self.get(label)]

    def get_ttm(self, label: str, use_pct: bool = False) -> Optional[float]:
        """TTM value for a label, or None if TTM column absent."""
        raw = self._ttm.get(label) or self._ttm.get(_norm(label))
        if raw is None:
            return None
        return _pct(raw) if use_pct else _f(raw)

    def sparse_series(self, label: str, use_pct: bool = False) -> list:
        """Non-None values only. Use only when gaps mean 'data does not exist'."""
        vals = self.get_pct(label) if use_pct else self.get_floats(label)
        return [v for v in vals if v is not None]

    def get_latest(self, label: str, use_pct: bool = False) -> Optional[float]:
        """Most recent non-None value."""
        vals = self.get_pct(label) if use_pct else self.get_floats(label)
        for v in reversed(vals):
            if v is not None:
                return v
        return None

    def last_n(self, label: str, n: int = 4, use_pct: bool = False) -> list:
        """Last N periods as [(date_label, value)] — for recent trend."""
        vals  = self.get_pct(label) if use_pct else self.get_floats(label)
        pairs = list(zip(self.dates, vals))
        return [(d, v) for d, v in pairs[-n:] if v is not None]

    def annual_years(self) -> list:
        """Mar-2024 → FY24. Other months returned as-is."""
        years = []
        for d in self.dates:
            if 'Mar' in d:
                yr = d.split('-')[1]
                years.append(f"FY{yr[2:]}")
            else:
                years.append(d)
        return years


# ─── Sheet parser ──────────────────────────────────────────────────────────────

SECTION_NAMES = {
    'Quarterly Results',
    'Annual P&L',
    'Annual P&L ',       # trailing space variant
    'Balance Sheet',
    'Cashflow statement',
    'Ratios',
    'Shareholding Pattern',
    'Links',
}

META_FIELDS = {
    'TICKER SYMBOL': 'ticker',
    'COMPANY NAME':  'company_name',
    'EXCHANGE':      'exchange',
    'INDUSTRY':      'industry',
    'SECTOR':        'sector',
    'DESCRIPTION':   'description',
    'CURRENCY':      'currency',
}


def _finalise_section(name, start, end, rows, sections, label_col, data_start):
    section_rows = rows[start:end]
    date_row     = None
    date_row_idx = None

    for j, row in enumerate(section_rows):
        row = list(row) + [None] * 15
        # Count datetime objects AND check for TTM string
        date_count = sum(1 for v in row[data_start:] if isinstance(v, datetime.datetime))
        if date_count >= 2:
            date_row     = row
            date_row_idx = j
            break

    if date_row is None:
        return

    canonical = name.strip()
    sections[canonical] = SheetSection(
        canonical,
        date_row,
        section_rows[date_row_idx + 1:],
        label_col=label_col,
        data_start=data_start,
    )


def _parse_sheet(rows: list, label_col: int = 0, data_start: int = 1) -> dict:
    meta     = {}
    sections = {}
    links    = {'transcripts': [], 'annual_reports': []}

    current_section       = None
    current_section_start = None
    link_mode             = None

    for i, row in enumerate(rows):
        row   = list(row) + [None] * 15
        label = _s(row[label_col])
        val   = _s(row[data_start])

        # Meta
        if label and label.upper() in META_FIELDS:
            meta[META_FIELDS[label.upper()]] = val
            continue

        # Section header
        matched = None
        if label:
            for sname in SECTION_NAMES:
                if label.strip() == sname.strip():
                    matched = sname.strip()
                    break
        if matched:
            if current_section and current_section != 'Links':
                _finalise_section(
                    current_section, current_section_start,
                    i, rows, sections, label_col, data_start
                )
            current_section       = matched
            current_section_start = i
            link_mode             = None
            continue

        # Links
        if current_section == 'Links':
            if label == 'Transcripts':
                link_mode = 'transcripts'
            elif label == 'Annual Reports':
                link_mode = 'annual_reports'
            for v in row:
                sv = _s(v)
                if sv and sv.startswith('http') and link_mode:
                    links[link_mode].append(sv)

    if current_section and current_section != 'Links':
        _finalise_section(
            current_section, current_section_start,
            len(rows), rows, sections, label_col, data_start
        )

    return {'meta': meta, 'sections': sections, 'links': links}


# ─── Annualisation fallback ────────────────────────────────────────────────────

def _annualise(quarterly: list) -> list:
    """Sum groups of 4 quarters → annual. Fallback only."""
    q = [v for v in quarterly if v is not None]
    if len(q) < 2:
        return q
    result = []
    i = 0
    while i + 4 <= len(q):
        result.append(round(sum(q[i:i + 4]), 2))
        i += 4
    remainder = q[i:]
    if remainder and not result:
        result.append(round(sum(remainder), 2))
    return result


def _avg_groups(vals: list, group_size: int = 4) -> list:
    clean = [v for v in vals if v is not None]
    result = []
    i = 0
    while i + group_size <= len(clean):
        chunk = clean[i:i + group_size]
        result.append(round(sum(chunk) / len(chunk), 2))
        i += group_size
    return result


# ─── PrimaryData builder ───────────────────────────────────────────────────────

def _build_from_parsed(
    parsed: dict,
    existing_holding:   Optional[bool]  = None,
    current_weight_pct: Optional[float] = None,
    avg_buy_price:      Optional[float] = None,
    original_thesis:    Optional[str]   = None,
) -> PrimaryData:

    raw_meta = parsed['meta']
    sections = parsed['sections']
    links    = parsed['links']

    # ── Meta ──
    exchange_str = (raw_meta.get('exchange') or 'OTHER').upper()
    try:
        exchange = Exchange(exchange_str)
    except ValueError:
        exchange = Exchange.OTHER

    meta = CompanyMeta(
        ticker      = raw_meta.get('ticker', 'UNKNOWN'),
        name        = raw_meta.get('company_name', 'Unknown'),
        exchange    = exchange,
        sector      = raw_meta.get('sector', raw_meta.get('industry', 'Unknown')),
        industry    = raw_meta.get('industry', 'Unknown'),
        currency    = raw_meta.get('currency', 'INR'),
        description = raw_meta.get('description'),
    )

    annual_pl = sections.get('Annual P&L')
    periodic  = sections.get('Quarterly Results')
    bs        = sections.get('Balance Sheet')
    cf        = sections.get('Cashflow statement')
    rt        = sections.get('Ratios')
    sh        = sections.get('Shareholding Pattern')

    reporting_frequency = periodic.frequency if periodic else 'unknown'

    # ──────────────────────────────────────────────────────────────────────────
    # INCOME
    # ──────────────────────────────────────────────────────────────────────────

    if annual_pl:
        annual_years = annual_pl.annual_years()

        revenue         = annual_pl.get_floats('Sales')
        ebitda          = annual_pl.get_floats('Operating Profit')
        ebitda_margin   = annual_pl.get_pct('OPM %')
        pat             = annual_pl.get_floats('Net Profit')
        eps             = annual_pl.get_floats('EPS in Rs')
        dividend_payout = annual_pl.get_pct('Dividend Payout %')

        # Revenue growth
        rev_growth = [None]
        for i in range(1, len(revenue)):
            prev, curr = revenue[i-1], revenue[i]
            if prev and prev != 0 and curr is not None:
                rev_growth.append(round((curr - prev) / abs(prev) * 100, 2))
            else:
                rev_growth.append(None)

        # PAT margin
        pat_margin = [
            round(p / r * 100, 2) if (p is not None and r and r != 0) else None
            for p, r in zip(pat, revenue)
        ]

        # TTM values — from TTM column if present
        ttm_revenue       = annual_pl.get_ttm('Sales')
        ttm_ebitda        = annual_pl.get_ttm('Operating Profit')
        ttm_ebitda_margin = annual_pl.get_ttm('OPM %', use_pct=True)
        ttm_pat           = annual_pl.get_ttm('Net Profit')
        ttm_eps           = annual_pl.get_ttm('EPS in Rs')

        income = IncomeStatement(
            revenue          = revenue,
            revenue_growth   = rev_growth,
            ebitda           = ebitda,
            ebitda_margin    = ebitda_margin,
            pat              = pat,
            pat_margin       = pat_margin,
            eps              = eps,
            dividend_payout  = dividend_payout,
            years            = annual_years,
            ttm_revenue      = ttm_revenue,
            ttm_ebitda       = ttm_ebitda,
            ttm_ebitda_margin= ttm_ebitda_margin,
            ttm_pat          = ttm_pat,
            ttm_eps          = ttm_eps,
            data_source      = 'Annual P&L',
        )

    elif periodic:
        # Fallback — aggregate from periodic
        group_size = 2 if reporting_frequency == 'semi-annual' else 4
        q_rev  = periodic.get_floats('Sales')
        q_pat  = periodic.get_floats('Net Profit')
        q_opm  = periodic.get_pct('OPM %')
        q_eps  = periodic.get_floats('EPS in Rs')

        annual_revenue = _annualise(q_rev) if group_size == 4 else q_rev
        annual_pat     = _annualise(q_pat) if group_size == 4 else q_pat
        annual_opm     = _avg_groups(q_opm, group_size)
        annual_eps     = _annualise(q_eps) if group_size == 4 else q_eps

        bs_years = bs.annual_years() if bs else []
        n = min(len(annual_revenue), len(bs_years) if bs_years else 999)
        eff_years = bs_years[:n] if bs_years else [f"Y{i+1}" for i in range(n)]

        rev_growth = [None]
        for i in range(1, len(annual_revenue)):
            prev, curr = annual_revenue[i-1], annual_revenue[i]
            if prev and prev != 0 and curr is not None:
                rev_growth.append(round((curr - prev) / abs(prev) * 100, 2))
            else:
                rev_growth.append(None)

        pat_margin = [
            round(p / r * 100, 2) if (p is not None and r and r != 0) else None
            for p, r in zip(annual_pat, annual_revenue)
        ]

        income = IncomeStatement(
            revenue        = annual_revenue[:n],
            revenue_growth = rev_growth[:n],
            ebitda         = [],
            ebitda_margin  = annual_opm[:n],
            pat            = annual_pat[:n],
            pat_margin     = pat_margin[:n],
            eps            = annual_eps[:n],
            years          = eff_years,
            data_source    = f'Aggregated from {reporting_frequency} data (no Annual P&L section)',
        )
    else:
        income = IncomeStatement(data_source='No income data found in sheet')

    # ── Recent trend — last 4 periods from periodic ──
    recent_periods = {}
    if periodic:
        trend_labels = ['Sales', 'OPM %', 'Net Profit', 'EPS in Rs']
        pct_labels   = {'OPM %'}
        for lbl in trend_labels:
            trend = periodic.last_n(lbl, n=4, use_pct=(lbl in pct_labels))
            if trend:
                recent_periods[lbl] = trend
        recent_periods['_frequency'] = reporting_frequency
        recent_periods['_periods']   = 4

    # ──────────────────────────────────────────────────────────────────────────
    # BALANCE SHEET
    # ──────────────────────────────────────────────────────────────────────────

    if bs:
        bs_years     = bs.annual_years()
        equity_cap   = bs.get_floats('Equity Capital')
        reserves     = bs.get_floats('Reserves')
        total_equity = [
            (e + r) if (e is not None and r is not None) else None
            for e, r in zip(equity_cap, reserves)
        ]
        total_debt        = bs.get_floats('Borrowings')
        long_term_debt    = bs.get_floats('Long term Borrowings')
        short_term_debt   = bs.get_floats('Short term Borrowings')
        total_assets      = bs.get_floats('Total Assets')
        cash              = bs.get_floats('Cash Equivalents')
        inventories       = bs.get_floats('Inventories')
        trade_receivables = bs.get_floats('Trade receivables')

        # ROCE from Ratios — preserve None for alignment
        roce = rt.get_pct('ROCE %') if rt else []

        debt_to_equity = [
            round(d / e, 3) if (d is not None and e and e != 0) else None
            for d, e in zip(total_debt, total_equity)
        ]
    else:
        bs_years = []
        total_equity = total_debt = long_term_debt = short_term_debt = []
        total_assets = cash = inventories = trade_receivables = roce = debt_to_equity = []

    balance = BalanceSheet(
        total_assets      = total_assets,
        total_equity      = total_equity,
        total_debt        = total_debt,
        cash              = cash,
        roce              = roce,
        roe               = [],
        debt_to_equity    = debt_to_equity,
        years             = bs_years,
        long_term_debt    = long_term_debt,
        short_term_debt   = short_term_debt,
        inventories       = inventories,
        trade_receivables = trade_receivables,
    )

    # ──────────────────────────────────────────────────────────────────────────
    # CASHFLOW
    # ──────────────────────────────────────────────────────────────────────────

    if cf:
        cf_years = cf.annual_years()
        cfo      = cf.get_floats('Cash from Operating Activity')
        cfi      = cf.get_floats('Cash from Investing Activity')
        cff      = cf.get_floats('Cash from Financing Activity')
        capex    = cf.get_floats('Fixed assets purchased')

        fcf = [
            round(o + k, 2) if (o is not None and k is not None) else None
            for o, k in zip(cfo, capex)
        ]
    else:
        cf_years = []
        cfo = cfi = cff = capex = fcf = []

    cashflow = CashflowStatement(
        cfo   = cfo,
        capex = capex,
        fcf   = fcf,
        years = cf_years,
        cfi   = cfi,
        cff   = cff,
    )

    # ──────────────────────────────────────────────────────────────────────────
    # OPERATING METRICS (from Ratios section)
    # ──────────────────────────────────────────────────────────────────────────

    if rt:
        rt_years = rt.annual_years()
        operating_metrics = OperatingMetrics(
            debtor_days           = rt.get_floats('Debtor Days'),
            inventory_days        = rt.get_floats('Inventory Days'),
            days_payable          = rt.get_floats('Days Payable'),
            cash_conversion_cycle = rt.get_floats('Cash Conversion Cycle'),
            working_capital_days  = rt.get_floats('Working Capital Days'),
            years                 = rt_years,
        )
    else:
        operating_metrics = OperatingMetrics()

    # ──────────────────────────────────────────────────────────────────────────
    # SHAREHOLDING
    # ──────────────────────────────────────────────────────────────────────────

    if sh:
        shareholding = ShareholdingData(
            promoters  = sh.get_floats('Promoters'),
            fiis       = sh.get_floats('FIIs'),
            diis       = sh.get_floats('DIIs'),
            government = sh.get_floats('Government'),
            public     = sh.get_floats('Public'),
            quarters   = sh.dates,
        )
    else:
        shareholding = ShareholdingData()

    # ──────────────────────────────────────────────────────────────────────────
    # RISK FLAGS
    # ──────────────────────────────────────────────────────────────────────────

    risk_notes = []

    # Promoter holding trend
    if sh and shareholding.promoters:
        ph = [v for v in shareholding.promoters if v is not None]
        if len(ph) >= 2 and ph[-1] < ph[-2]:
            risk_notes.append(
                f"PROMOTER_SELLING: Promoter holding declined from "
                f"{ph[-2]*100:.1f}% to {ph[-1]*100:.1f}%"
            )

    # Institutional exit — both DII and FII declining
    if sh:
        diis = [v for v in shareholding.diis if v is not None]
        fiis = [v for v in shareholding.fiis if v is not None]
        if len(diis) >= 2 and diis[-1] < diis[-2]:
            if len(fiis) >= 2 and fiis[-1] < fiis[-2]:
                risk_notes.append(
                    "INSTITUTIONAL_EXIT: Both DII and FII holdings declining"
                )

    # Debtor days expanding
    if rt and operating_metrics.debtor_days:
        dd = [v for v in operating_metrics.debtor_days if v is not None]
        if len(dd) >= 2 and dd[-1] > dd[-2] * 1.2:
            risk_notes.append(
                f"DEBTOR_DAYS_EXPANDING: {dd[-2]:.0f} → {dd[-1]:.0f} days "
                f"(+{((dd[-1]/dd[-2])-1)*100:.0f}%) — review collections"
            )

    # Debt rising
    if bs and balance.total_debt:
        debt = [v for v in balance.total_debt if v is not None]
        if len(debt) >= 2 and debt[-1] > debt[-2] * 1.3:
            risk_notes.append(
                f"DEBT_INCREASE: Total borrowings up "
                f"{((debt[-1]/debt[-2])-1)*100:.0f}% YoY"
            )

    # SME semi-annual flag
    if reporting_frequency == 'semi-annual':
        risk_notes.append(
            "SME_STOCK: Semi-annual reporting — "
            "recent trend window is half-years not quarters"
        )

    # Data quality
    if not annual_pl:
        risk_notes.append(
            "DATA_QUALITY: Annual P&L section missing — "
            "income aggregated from periodic data (less reliable)"
        )

    # Transcripts
    if links['transcripts']:
        risk_notes.append(
            f"TRANSCRIPTS_AVAILABLE: {len(links['transcripts'])} PDFs "
            f"(feed to RAG in P5)"
        )

    risk = RiskMeta(
        fx_exposure        = None,
        related_party_flag = None,
        notes              = risk_notes,
    )

    # ──────────────────────────────────────────────────────────────────────────
    # ASSEMBLE
    # ──────────────────────────────────────────────────────────────────────────

    primary = PrimaryData(
        meta               = meta,
        income             = income,
        balance            = balance,
        cashflow           = cashflow,
        valuation          = ValuationInputs(),
        risk_meta          = risk,
        shareholding       = shareholding,
        operating_metrics  = operating_metrics,
        existing_holding   = existing_holding,
        current_weight_pct = current_weight_pct,
        avg_buy_price      = avg_buy_price,
        original_thesis    = original_thesis,
        recent_periods     = recent_periods,
        reporting_frequency= reporting_frequency,
    )
    primary.data_fingerprint = primary.compute_fingerprint()
    return primary


# ─── Public entry points ───────────────────────────────────────────────────────

def load_from_xlsx(path: str, ticker: str = None, **position_kwargs) -> PrimaryData:
    """
    Load from local .xlsx file.
    V1 contract: labels in col 0, data from col 1.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)

    if ticker and ticker.upper() in wb.sheetnames:
        ws = wb[ticker.upper()]
    elif len(wb.sheetnames) == 1:
        ws = wb[wb.sheetnames[0]]
    elif ticker:
        matches = [s for s in wb.sheetnames if ticker.upper() in s.upper()]
        ws = wb[matches[0]] if matches else wb.active
    else:
        ws = wb.active

    rows   = list(ws.iter_rows(values_only=True))
    parsed = _parse_sheet(rows, label_col=0, data_start=1)
    return _build_from_parsed(parsed, **position_kwargs)


def load_from_google_sheets(
    sheet_id:    str,
    ticker:      str,
    sa_key_path: Optional[str] = None,
    **position_kwargs
) -> PrimaryData:
    """
    Load from Google Sheets. Tab must match ticker symbol.
    gspread: labels col 0, data from col 1.
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        raise ImportError("pip install gspread google-auth")

    key_path = sa_key_path or os.environ.get("GOOGLE_SA_KEY")
    if not key_path:
        raise ValueError("Set GOOGLE_SA_KEY env var or pass sa_key_path")

    creds = Credentials.from_service_account_file(
        key_path,
        scopes=[
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
    )
    client = gspread.authorize(creds)
    sheet  = client.open_by_key(sheet_id)

    try:
        ws = sheet.worksheet(ticker.upper())
    except Exception:
        ws = sheet.sheet1

    raw_rows  = ws.get_all_values()
    converted = [tuple(_parse_date(cell) for cell in row) for row in raw_rows]
    parsed    = _parse_sheet(converted, label_col=0, data_start=1)
    return _build_from_parsed(parsed, **position_kwargs)


def load_from_dict(data: dict, **position_kwargs) -> PrimaryData:
    """For unit tests."""
    return _build_from_parsed(data, **position_kwargs)
